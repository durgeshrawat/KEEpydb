from msilib.schema import Error
import openpyxl
import pyrebase
from openpyxl import load_workbook        
import os
import urllib
import json
#CORE CLASS FOR KEEpydb DATABASE
class core:
    def create_database(self,dbname,username,password):
        self.db=openpyxl.Workbook()
        os.system('mkdir '+dbname)
        try:
            self.db.save(f'{dbname}/{username}.KEEpydb.xlsx')
            with open(f'{dbname}/dbsetups.KEEpydb','w') as f:
                f.write('0')
            with open(f'{dbname}/__configs__','w') as g:
                data=f'status=active,{username},{password}'
                g.write(data)
            return True
        except:
            self.db.save(f'{dbname}\\{username}.KEEpydb.xlsx')
            with open(f'{dbname}\\dbsetups.KEEpydb','w') as f:
                f.write('0')
            with open(f'{dbname}\\__configs__','w') as g:
                data=f'status=active,{username},{password}'
                g.write(data)
            return True
            
class tools:
    def CE(self,dbname,username,password):
        with open(f'{dbname}/dbsetups.KEEpydb','r') as f:
            ff=f.readlines()
            ff=ff[0]
            try:
                if ff == '0':
                    os.rename(f'{dbname}/{username}.KEEpydb',f'{dbname}/{username}.KEEpydb.xlsx')
                    with open(f'{dbname}/dbsetups.KEEpydb','w') as g:
                        g.write('1')
                elif ff == '1':
                    os.rename(f'{dbname}/{username}.KEEpydb.xlsx',f'{dbname}/{username}.KEEpydb')
                    with open(f'{dbname}/dbsetups.KEEpydb','w') as g:
                        g.write('0')
            except:
                if ff == '0':
                    os.rename(f'{dbname}\\{username}.KEEpydb',f'{dbname}\\{username}.KEEpydb.xlsx')
                    with open(f'{dbname}\\dbsetups.KEEpydb','w') as g:
                        g.write('1')
                elif ff == '1':
                    os.rename(f'{dbname}\\{username}.KEEpydb.xlsx',f'{dbname}\\{username}.KEEpydb')
                    with open(f'{dbname}\\dbsetups.KEEpydb','w') as g:
                        g.write('0')

class query:
    def __init__(self,dbname,username,password):
        self.dbname=dbname
        self.username=username
        self.password=password
        try:
            self.workbook = load_workbook(filename=f'{self.dbname}/{self.username}.KEEpydb.xlsx')
        except:
            self.workbook = load_workbook(filename=f'{self.dbname}\\{self.username}.KEEpydb.xlsx')
        self.sheet=self.workbook.active
        
    def update(self,columnANDrow,value): #append new data
        self.sheet[columnANDrow] = value

    def get_cell(self,cell_no): # get perticular cell
        return self.sheet[cell_no].value
        
    def get_all(self): #show whole database
        l=[]
        for i in self.sheet.iter_rows(values_only=True):
            l.append(i)        
        return l

    def get_from_columns(self,column): # example column = B
        l=[]
        for i in self.sheet[column]:
            l.append(i.value)
        return l

    def get_from_range_columns(self,column_range): # column_range is in form of str and as '2:6' means from 2 to 6
        l=[]
        for i in self.sheet[column_range]:
            l.append(i.value)
        return l
        
    def get_from_rows(self,row): #example row = 5
        l=[]
        for i in self.sheet[row]:
            l.append(i.value)
        return l

    def get_from_range_rows(self,ranged_row): # example of ranged row = '3:5'
        l=[]
        for i in self.sheet[ranged_row]:
            l.append(i.value)
        return l

    def add_columns(self,idx,amount=1): #idx = index no of column will be , amount = no of columns do you want to insert in database
        self.sheet.insert_cols(idx,amount)

    def add_rows(self,idx,amount=1): #insert rows
        self.sheet.insert_rows(idx,amount)

    def delete_cell(self,cell): # example cell = a2
        self.sheet[cell]=None

    def delete_columns(self,idx,amount=1): #delete columns
        self.sheet.delete_cols(idx,amount)
        
    def search(self,searchelement): #search for values
        l=[]
        for i in self.sheet.iter_rows(values_only=True):
            l.append(i)
        for i in l:
            for j in i:
                if searchelement in i:
                    return (searchelement,i[i.index(searchelement)+1])
                    
    def delete_rows(self,idx,amoumt=1): #delete row
        self.sheet.delete_rows(idx,amoumt)
        
    def save(self): #save update after all updation
        try:
            self.workbook.save(filename=f'{self.dbname}/{self.username}.KEEpydb.xlsx')
        except:
            self.workbook.save(filename=f'{self.dbname}\\{self.username}.KEEpydb.xlsx')
            
class realtimedatabase:
    def __init__(self,authtoken):
        try:
            firebase = pyrebase.initialize_app(authtoken)
            self.auth= firebase.auth()
            self.storage=firebase.storage()
        except Exception as e:
            print(e)

    def login(self,email,password):
        try:
            info=self.auth.sign_in_with_email_and_password(email=email, password=password,)
            if info["registered"]==True:
                return True
        except Exception as getaddrinfo :
            raise Error(getaddrinfo)
    
    def adduser(self,email,password,name="KEEpydb-Client"):
        try:
            print(self.auth.create_user_with_email_and_password(email,password))
        except Exception as getaddrinfo :
            raise Error(getaddrinfo)

    #Storage
    def upload(self,filename,cloudfilename,returnurl=False):
        self.storage.child(cloudfilename).put(filename)
        if returnurl==True:
            return self.storage.child(cloudfilename).get_url(None)
    
    def geturl(self,cloudfilename):
        return self.storage.child(cloudfilename).get_url(None)


    def download(self,cloudfilename,path,filename):
        self.storage.child(cloudfilename).download(path,filename)

    def readfile(self,cloudfilename,decode=True,file=False):
        global cloudfilename1
        if file==False:
            cloudfilename1=''
            for j in str(cloudfilename):
                if j == '.':
                    cloudfilename1+='/'
                else:
                    cloudfilename1+=j

            url=self.geturl(cloudfilename1)
            f=urllib.request.urlopen(url).read()
            if decode==True:
                return f.decode()
            else:
                return f
        else:
            url=self.geturl(cloudfilename)
            f=urllib.request.urlopen(url).read()
            if decode==True:
                return f.decode()

    def isobject(self,objectname):
        try:
            url=self.geturl(objectname)
            f=urllib.request.urlopen(url).read()
            return True
        except:
            return False
    def createobject(self,objectname):
        objectname=str(objectname+"/ObjectHandle")
        with open("objectHandle","w") as f:
            f.write("This is an <KEEpydbObject>")
        self.upload("objectHandle",objectname)
        os.system("del objectHandle")
        return True
    def createfile(self,filename,dict):
        with open("objectHandle","w") as f:
            f.write(json.dumps(dict))
        self.upload("objectHandle",filename)
        os.system("del objectHandle")
        return True

    def pushdata(self,varname,data,file=False): #fileobjectclassname[object class name],variablename[filename],data
        global varname1
        varname1=''                  # data store in .json format {:}
        if file == False:
            for j in varname:
                if j == '.':
                    varname1+='/'
                else:
                    varname1+=j
            with open("objectfile","w") as f:
                f.write(json.dumps(data))
            self.upload('objectfile',varname1,data)     #filename,cloudfilename
            os.system("del objectfile")
        else:
            with open("objectfile","w") as f:
                f.write(json.dumps(data))
            self.upload('objectfile',varname,data)     #filename,cloudfilename
            os.system("del objectfile")
        return True
        
    def objectcall(self,objectname,file=False):
        if file==True:
            mydata=self.readfile(objectname,file=True)
            return json.loads(mydata)

        mydata=self.readfile(objectname)
        return json.loads(mydata)
    
    def recordappend(self,objectname,data):#objectname fileaddress in db ; data a dictionary
        mydata=self.objectcall(objectname,file=True)
        for i in mydata['records']:
            if i == data['records']:
                return False
        mydata['records'].append(data['records'])
        self.pushdata(objectname,mydata)
        return True
        

